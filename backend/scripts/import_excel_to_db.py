#!/usr/bin/env python3
"""Import components from an Excel file into SQLite."""

import argparse
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

DB_COLUMNS = [
    "nom",
    "reference",
    "categorie",
    "description",
    "quantite",
    "emplacement",
    "prix",
    "photo_url",
]
REQUIRED_COLUMNS = ["nom", "reference"]
TEXT_COLUMNS = ["nom", "reference", "categorie", "description", "emplacement", "photo_url"]


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def optional_text(value: Any) -> Optional[str]:
    text = normalize_text(value)
    return text if text else None


def parse_int(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    text = normalize_text(value).replace(" ", "").replace(",", ".")
    if not text:
        return 0
    return int(round(float(text)))


def parse_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value)
    text = text.replace(" ", "").replace("€", "").replace(",", ".")
    if not text:
        return 0.0
    return float(text)


def load_rows_from_excel(
    excel_path: Path, sheet_name: Optional[str]
) -> Tuple[List[Dict[str, Any]], int, int]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(value) for value in header_row]
    index_by_name = {name: idx for idx, name in enumerate(headers) if name}

    missing = [name for name in REQUIRED_COLUMNS if name not in index_by_name]
    if missing:
        missing_text = ", ".join(missing)
        raise SystemExit(f"Missing required columns: {missing_text}")

    rows_by_reference: Dict[str, Dict[str, Any]] = {}
    skipped = 0
    merged = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(value is None or str(value).strip() == "" for value in row):
            continue

        row_values: Dict[str, Any] = {}
        for column in DB_COLUMNS:
            idx = index_by_name.get(column)
            row_values[column] = row[idx] if idx is not None else None

        nom = normalize_text(row_values["nom"])
        reference = normalize_text(row_values["reference"])
        if not nom or not reference:
            skipped += 1
            continue

        item = {
            "nom": nom,
            "reference": reference,
            "categorie": optional_text(row_values["categorie"]),
            "description": optional_text(row_values["description"]),
            "quantite": parse_int(row_values["quantite"]),
            "emplacement": optional_text(row_values["emplacement"]),
            "prix": parse_float(row_values["prix"]),
            "photo_url": optional_text(row_values["photo_url"]),
        }

        existing = rows_by_reference.get(reference)
        if existing:
            merged += 1
            existing["quantite"] += item["quantite"]
            for column in TEXT_COLUMNS:
                if not existing[column] and item[column]:
                    existing[column] = item[column]
            if existing["prix"] == 0.0 and item["prix"]:
                existing["prix"] = item["prix"]
        else:
            rows_by_reference[reference] = item

    return list(rows_by_reference.values()), skipped, merged


def wipe_tables(conn: sqlite3.Connection, mode: str) -> None:
    if mode == "none":
        return

    if mode == "all":
        tables = ["bom_lignes", "mouvements_stock", "composants", "projets"]
    else:
        tables = ["bom_lignes", "mouvements_stock", "composants"]

    for table in tables:
        conn.execute(f"DELETE FROM {table}")

    placeholders = ",".join(["?"] * len(tables))
    conn.execute(
        f"DELETE FROM sqlite_sequence WHERE name IN ({placeholders})",
        tables,
    )


def insert_components(conn: sqlite3.Connection, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        return

    sql = (
        "INSERT INTO composants (nom, reference, categorie, description, quantite, emplacement, prix, photo_url) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?) "
        "ON CONFLICT(reference) DO UPDATE SET "
        "quantite = composants.quantite + excluded.quantite, "
        "nom = COALESCE(composants.nom, excluded.nom), "
        "categorie = COALESCE(composants.categorie, excluded.categorie), "
        "description = COALESCE(composants.description, excluded.description), "
        "emplacement = COALESCE(composants.emplacement, excluded.emplacement), "
        "prix = CASE WHEN composants.prix = 0 THEN excluded.prix ELSE composants.prix END, "
        "photo_url = COALESCE(composants.photo_url, excluded.photo_url)"
    )

    values = [
        (
            row["nom"],
            row["reference"],
            row["categorie"],
            row["description"],
            row["quantite"],
            row["emplacement"],
            row["prix"],
            row["photo_url"],
        )
        for row in rows
    ]

    conn.executemany(sql, values)


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    backend_dir = script_dir.parent
    repo_dir = backend_dir.parent

    parser = argparse.ArgumentParser(description="Import Excel components into atelier.db")
    parser.add_argument(
        "--excel",
        default=str(repo_dir / "Inventaire_Electronique.xlsx"),
        help="Path to Excel file (default: Inventaire_Electronique.xlsx in repo root)",
    )
    parser.add_argument(
        "--db",
        default=str(backend_dir / "data" / "atelier.db"),
        help="Path to SQLite database (default: backend/data/atelier.db)",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name to import (default: active sheet)",
    )
    parser.add_argument(
        "--wipe",
        choices=["all", "components", "none"],
        default="none",
        help="Delete existing rows before import (default: none)",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    excel_path = Path(args.excel).resolve()
    db_path = Path(args.db).resolve()

    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")
    if not db_path.exists():
        raise SystemExit(f"Database file not found: {db_path}")

    rows, skipped, merged = load_rows_from_excel(excel_path, args.sheet)

    conn = sqlite3.connect(db_path)
    try:
        conn.execute("PRAGMA foreign_keys = ON")
        with conn:
            wipe_tables(conn, args.wipe)
            insert_components(conn, rows)
    finally:
        conn.close()

    print(f"Imported {len(rows)} components into {db_path}")
    if skipped:
        print(f"Skipped {skipped} rows missing nom/reference")
    if merged:
        print(f"Merged {merged} duplicate references")


if __name__ == "__main__":
    main()
